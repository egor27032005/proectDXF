from flask import Flask, request, send_file
import io
import ezdxf
import os
import logging
import warnings
from waitress import serve
from openpyxl import load_workbook  # Импортируем openpyxl

from flask_cors import CORS

from PythonFiles.FullPreparation.StartPreparation import StartPreparation

# Игнорируем предупреждение о Data Validation
warnings.filterwarnings("ignore", category=UserWarning, message="Data Validation extension is not supported and will be removed")

app = Flask(__name__)
CORS(app)

# Настройка логирования
logging.basicConfig(level=logging.ERROR)
@app.route('/')
def health_check():
    return "OK", 200

@app.route('/process-excel', methods=['POST'])
def process_excel():
    if 'file' not in request.files or 'option' not in request.form:
        return {"error": "No file or option provided"}, 400
    file = request.files['file']
    option = request.form['option']
    if file.filename == '':
        return {"error": "No file selected"}, 400
    if not file.filename.endswith('.xlsx') and not file.filename.endswith('.xls'):
        return {"error": "Unsupported file format. Only .xlsx and .xls are supported"}, 400
    try:
        temp_excel_file = "temp_input.xlsx"
        file.save(temp_excel_file)
        wb = load_workbook(filename=temp_excel_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        os.remove(temp_excel_file)
        doc = ezdxf.new("R2000")
        msp = doc.modelspace()
        st = StartPreparation(data, option, msp, doc)
        temp_dxf_file = "temp_output.dxf"
        doc.saveas(temp_dxf_file)
        with open(temp_dxf_file, "rb") as f:
            dxf_stream = io.BytesIO(f.read())
        os.remove(temp_dxf_file)
        dxf_stream.seek(0)
        return send_file(
            dxf_stream,
            mimetype='application/octet-stream',
            as_attachment=True,
            download_name='output.dxf'
        )

    except Exception as e:
        app.logger.error(f"Error processing file: {e}")
        return {"error": "Failed to process the file"}, 500


if __name__ == '__main__':
    # Запуск через Waitress для production
    serve(app, host="0.0.0.0", port=5000)