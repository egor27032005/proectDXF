from flask import Flask, request, send_file, jsonify
import io
import ezdxf
import os
import logging
import warnings
import sys
from waitress import serve
from openpyxl import load_workbook
from flask_cors import CORS

# Импорт вашего класса обработки (замените на реальный путь)
from PythonFiles.FullPreparation.StartPreparation import StartPreparation

# Игнорируем предупреждение о Data Validation
warnings.filterwarnings("ignore", category=UserWarning,
                        message="Data Validation extension is not supported and will be removed")

app = Flask(__name__)
CORS(app)

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('app.log')
    ]
)
logger = logging.getLogger(__name__)


@app.route('/')
def health_check():
    return "OK", 200


def excel_to_dict(file_path, read_only=True):
    """
    Конвертирует Excel файл в словарь, где ключи - названия листов,
    а значения - данные из этих листов (списки списков)

    Args:
        file_path (str): Путь к файлу Excel
        read_only (bool): Режим только для чтения (оптимизация для больших файлов)

    Returns:
        dict: Словарь с данными всех листов
    """
    try:
        wb = load_workbook(filename=file_path, read_only=read_only)
        sheets_data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = []

            logger.info(f"Processing sheet '{sheet_name}' ({ws.max_row} rows x {ws.max_column} cols)")

            if ws.max_row == 0:
                logger.warning(f"Sheet '{sheet_name}' is empty")
                sheets_data[sheet_name] = []
                continue

            for row in ws.iter_rows(values_only=True):
                sheet_data.append(list(row))

            sheets_data[sheet_name] = sheet_data

        return sheets_data

    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
        raise


@app.route('/process-excel', methods=['POST'])
def process_excel():
    """
    Основной endpoint для обработки Excel файла и генерации DXF
    """
    if 'file' not in request.files or 'option' not in request.form:
        return jsonify({"error": "No file or option provided"}), 400

    file = request.files['file']
    option = request.form['option']

    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Unsupported file format. Only .xlsx and .xls are supported"}), 400

    # Проверка доступности директории для временных файлов
    if not os.access('.', os.W_OK):
        logger.error("No write permissions in current directory")
        return jsonify({"error": "Server configuration error"}), 500

    temp_excel_file = "temp_input.xlsx"
    temp_dxf_file = "temp_output.dxf"

    try:
        # Сохраняем временный файл
        file.save(temp_excel_file)
        logger.info(f"File saved temporarily as {temp_excel_file}")

        # Конвертируем Excel в словарь
        sheets_data = excel_to_dict(temp_excel_file)
        logger.info(f"Successfully loaded {len(sheets_data)} sheets")

        # Создаем DXF документ
        logger.info(f"Creating DXF document for option: {option}")
        doc = ezdxf.new("R2000")
        msp = doc.modelspace()

        # Обработка данных
        logger.info("Starting data processing...")
        # st = StartPreparation(sheets_data, option, msp, doc)
        logger.info("Data processing completed")

        # Сохраняем DXF
        doc.saveas(temp_dxf_file)
        logger.info(f"DXF saved temporarily as {temp_dxf_file}")

        # Читаем DXF файл в поток
        with open(temp_dxf_file, "rb") as f:
            dxf_stream = io.BytesIO(f.read())

        # Удаляем временные файлы
        os.remove(temp_excel_file)
        os.remove(temp_dxf_file)
        logger.info("Temporary files removed")

        # Возвращаем DXF файл
        dxf_stream.seek(0)
        return send_file(
            dxf_stream,
            mimetype='application/octet-stream',
            as_attachment=True,
            download_name=f'output_{option}.dxf'
        )

    except Exception as e:
        logger.error(f"Error processing file: {str(e)}", exc_info=True)

        # Очистка временных файлов в случае ошибки
        if os.path.exists(temp_excel_file):
            os.remove(temp_excel_file)
        if os.path.exists(temp_dxf_file):
            os.remove(temp_dxf_file)

        return jsonify({
            "error": "Failed to process the file",
            "details": str(e)
        }), 500


@app.route('/inspect', methods=['POST'])
def inspect_file():
    """
    Endpoint для проверки структуры файла без генерации DXF
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    temp_file = "temp_inspect.xlsx"

    try:
        file.save(temp_file)
        sheets_data = excel_to_dict(temp_file, read_only=True)

        response = {
            "filename": file.filename,
            "sheet_names": list(sheets_data.keys()),
            "sheets_preview": {
                name: data[:3] for name, data in sheets_data.items()
            },
            "sheets_count": len(sheets_data),
            "system_info": {
                "python_version": sys.version,
                "platform": sys.platform,
                "encoding": sys.getfilesystemencoding()
            }
        }

        os.remove(temp_file)
        return jsonify(response)

    except Exception as e:
        logger.error(f"Inspection error: {str(e)}", exc_info=True)
        if os.path.exists(temp_file):
            os.remove(temp_file)
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    logger.info("Starting server...")
    logger.info(f"Python version: {sys.version}")
    logger.info(f"System encoding: {sys.getfilesystemencoding()}")
    logger.info(f"Current directory: {os.getcwd()}")

    # Запуск через Waitress для production
    serve(app, host="0.0.0.0", port=5000)